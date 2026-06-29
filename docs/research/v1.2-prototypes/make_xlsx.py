#!/usr/bin/env python3
# Build the v1.2 scenario-grade workbook. Reproducible: run from anywhere.
#   python3 docs/research/v1.2-prototypes/make_xlsx.py
# Every quiz-answer scenario (Political x DEI x Animals x Guns x Priority-issue)
# x top companies -> grade under the v1.2 start-at-A prototype. Color-coded, filterable.
import json, os, itertools
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT=os.path.abspath(os.path.join(os.path.dirname(__file__),"..","..",".."))
DIR=os.path.join(ROOT,"public","data","companies")
OUT=os.path.join(ROOT,"docs","research","TruNorth-v1.2-scenario-grades.xlsx")

COMPANIES=[("patagonia","Patagonia"),("nike","Nike"),("apple","Apple"),("amazon","Amazon"),
 ("walmart","Walmart"),("costco","Costco"),("target","Target"),("starbucks","Starbucks"),
 ("chick-fil-a","Chick-fil-A"),("ben-and-jerry-s","Ben & Jerry's"),("exxonmobil","ExxonMobil"),
 ("chevron","Chevron"),("duke-energy","Duke Energy"),("general-motors","GM"),("coca-cola","Coca-Cola"),
 ("hershey","Hershey"),("dick-s-sporting-goods","Dick's"),("hobby-lobby","Hobby Lobby"),
 ("mcdonald-s","McDonald's"),("tesla","Tesla"),("meta-platforms","Meta"),("microsoft","Microsoft"),
 ("home-depot","Home Depot"),("disney","Disney"),("netflix","Netflix"),("procter-and-gamble","P&G"),
 ("johnson-and-johnson","J&J"),("delta-air-lines","Delta"),("bank-of-america","Bank of America"),("verizon","Verizon")]

BENCH=48; GOOD=78; Wd=[1,0.6,0.4,0.25,0.15]
clamp=lambda n,a,b: max(a,min(b,n))
def grade(n): return "?" if n is None else "A" if n>=88 else "B" if n>=70 else "C" if n>=52 else "D" if n>=36 else "F"
def live_grade(n): return "?" if n is None else "A" if n>=62 else "B" if n>=50 else "C" if n>=38 else "D" if n>=33 else "F"

def pol_score(lean,d):
    if not d: return None
    l=(lean or ""); l="left" if "left" in l else "right" if "right" in l else "bipartisan" if "bipartisan" in l else "neutral"
    return None if l=="neutral" else {"left":90 if d=="prog" else 22,"right":22 if d=="prog" else 90,"bipartisan":62}[l]
def dei_score(d,dr):
    if dr not in ("pro","anti"): return None
    s=d or ""; pro="pro_dei" in s; mix="mixed" in s; anti="anti" in s
    if dr=="pro": return 90 if pro else 55 if mix else 20 if anti else None
    return 22 if pro else 48 if mix else 90 if anti else None
def ani_score(a,care):
    if not care: return None
    s=a or ""
    if any(k in s for k in("positive","cruelty_free","excellent","good")): return 92
    if "mixed" in s: return 52
    if any(k in s for k in("negative","poor","tests","bad")): return 18
    return None
def gun_score(g,st):
    if not st: return None
    s=g or ""; sells=("sells_guns" in s) or ("makes_guns" in s); no="no_guns" in s
    if st=="avoid": return 18 if sells else 85 if no else None
    if st=="support": return 85 if sells else None
    return None

def compute(csc,sc,pol,dei,ani,gun,weights=None):
    weights=weights or {}
    cats=[(v,weights.get(k,1.0)) for k,v in (csc or {}).items()]
    for s,w in ((pol_score(sc.get("political"),pol),1.5),(dei_score(sc.get("dei"),dei),1.5),
                (ani_score(sc.get("animals"),ani),1.2),(gun_score(sc.get("guns"),gun),1.2)):
        if s is not None: cats.append((s,w))
    if not cats: return None
    harms=sorted([max(0,BENCH-s)*w for s,w in cats if max(0,BENCH-s)>0],reverse=True)
    penalty=sum(h*(Wd[i] if i<len(Wd) else 0.1) for i,h in enumerate(harms))
    credit=min(10,sum(1 for s,w in cats if s>=GOOD)*3)
    return clamp(100-penalty+credit,0,100)

cos=[{"name":nm,"sc":(json.load(open(f"{DIR}/{sl}.json")).get("sc",{})),
      "csc":(json.load(open(f"{DIR}/{sl}.json")).get("csc",{})),
      "overall":(json.load(open(f"{DIR}/{sl}.json")).get("overall"))} for sl,nm in COMPANIES]

FILL={"A":"54B98A","B":"A8D8B0","C":"E8C04C","D":"E8A04C","F":"E0524D","?":"BBBBBB"}
gfill=lambda g: PatternFill("solid",fgColor=FILL.get(g,"FFFFFF"))
HEAD=PatternFill("solid",fgColor="0E0F12"); HF=Font(color="EDE9E0",bold=True,size=9)
ANSF=Font(bold=True,size=9); CEN=Alignment(horizontal="center",vertical="center")
thin=Side(style="thin",color="DDDDDD"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
ROT=Alignment(text_rotation=90,horizontal="center",vertical="bottom")

wb=openpyxl.Workbook()
NA=len(COMPANIES)

# ---- Sheet 1: All Scenarios ----
ws=wb.active; ws.title="All Scenarios"
POL=[("Progressive","prog"),("Neutral",None),("Conservative","cons")]
DEI=[("Value DEI","pro"),("Neutral",None),("Avoid DEI","anti")]
ANI=[("Care",True),("Neutral",False)]
GUN=[("Avoid","avoid"),("Neutral",None),("Support","support")]
PRI=[("—",None),("Environment","environment"),("Labor","labor"),("Charity","charity"),("Privacy","privacy"),("Exec pay","execPay"),("Governance","governance")]
heads=["Political","DEI","Animals","Guns","Priority issue"]
for j,h in enumerate(heads,1):
    c=ws.cell(1,j,h); c.fill=HEAD; c.font=HF; c.alignment=CEN; c.border=BORD
for k,(sl,nm) in enumerate(COMPANIES):
    c=ws.cell(1,6+k,nm); c.fill=HEAD; c.font=HF; c.alignment=ROT; c.border=BORD
r=2
for pol,dei,ani,gun,pri in itertools.product(POL,DEI,ANI,GUN,PRI):
    for j,val in enumerate([pol[0],dei[0],ani[0],gun[0],pri[0]],1):
        cc=ws.cell(r,j,val); cc.font=Font(size=9); cc.alignment=Alignment(horizontal="center"); cc.border=BORD
    w={pri[1]:2.5} if pri[1] else {}
    for k,co in enumerate(cos):
        g=grade(compute(co["csc"],co["sc"],pol[1],dei[1],ani[1],gun[1],w))
        c=ws.cell(r,6+k,g); c.fill=gfill(g); c.alignment=CEN; c.border=BORD; c.font=Font(bold=True,size=10)
    r+=1
ws.freeze_panes="F2"; ws.auto_filter.ref=f"A1:{get_column_letter(5+NA)}{r-1}"
ws.row_dimensions[1].height=96
for j in range(1,6): ws.column_dimensions[get_column_letter(j)].width=[12,11,9,9,14][j-1]
for k in range(NA): ws.column_dimensions[get_column_letter(6+k)].width=5.2

# ---- Sheet 2: Named Profiles (incl. category-weight emphasis) ----
ws2=wb.create_sheet("Named Profiles")
PERS=[("Base — no quiz","neutral on everything",None,None,False,None,{}),
 ("Climate progressive","progressive · value DEI · care animals · avoid guns · env×2.2 labor×1.4","prog","pro",True,"avoid",{"environment":2.2,"labor":1.4}),
 ("Faith conservative","conservative · support guns · charity×1.6","cons",None,False,"support",{"charity":1.6}),
 ("Worker-first (union)","neutral stances · labor×2.5 execPay×1.6",None,None,False,None,{"labor":2.5,"execPay":1.6}),
 ("Privacy hawk","neutral stances · privacy×2.8",None,None,False,None,{"privacy":2.8}),
 ("Animal advocate","care animals · avoid guns · animals×2.5 env×1.3",None,None,True,"avoid",{"environment":1.3}),
 ("Conservative anti-DEI","conservative · AVOID DEI · neutral else","cons","anti",False,None,{})]
ws2.cell(1,1,"Profile").fill=HEAD; ws2.cell(1,1).font=HF
ws2.cell(1,2,"Quiz answers").fill=HEAD; ws2.cell(1,2).font=HF
for k,(sl,nm) in enumerate(COMPANIES):
    c=ws2.cell(1,3+k,nm); c.fill=HEAD; c.font=HF; c.alignment=ROT
# current live row
ws2.cell(2,1,"CURRENT (live, no quiz)").font=Font(bold=True,italic=True,size=9); ws2.cell(2,2,"the live app grade today").font=Font(size=9,italic=True)
for k,co in enumerate(cos):
    g=live_grade(co["overall"]); c=ws2.cell(2,3+k,g); c.fill=gfill(g); c.alignment=CEN; c.font=Font(bold=True,italic=True); c.border=BORD
for i,(nm,desc,pol,dei,ani,gun,w) in enumerate(PERS,3):
    ws2.cell(i,1,nm).font=ANSF; ws2.cell(i,2,desc).font=Font(size=9,italic=True)
    for k,co in enumerate(cos):
        g=grade(compute(co["csc"],co["sc"],pol,dei,ani,gun,w))
        c=ws2.cell(i,3+k,g); c.fill=gfill(g); c.alignment=CEN; c.font=Font(bold=True); c.border=BORD
ws2.freeze_panes="C2"; ws2.row_dimensions[1].height=96
ws2.column_dimensions["A"].width=22; ws2.column_dimensions["B"].width=52
for k in range(NA): ws2.column_dimensions[get_column_letter(3+k)].width=5.2

# ---- Sheet 3: Legend ----
ws3=wb.create_sheet("Legend & Method")
rows=[("TruNorth — v1.2 'start-at-A' scenario grades",""),("",""),
 ("PROTOTYPE — not live app grades. Models the proposed v1.2 model.",""),("",""),
 ("Grade colors:",""),("  A","Clean record — no documented harm found"),("  B","Minor / isolated issues"),
 ("  C","Moderate documented harm"),("  D","Serious documented harm"),("  F","Severe / multiple harms"),("  ?","Insufficient data"),("",""),
 ("Tab 'All Scenarios':","every combo of Political x DEI x Animals x Guns x Priority-issue (216 rows)."),
 ("","Use the filter arrows in row 1 to slice (e.g. Conservative + Avoid DEI)."),
 ("Tab 'Named Profiles':","7 ready-made personas — these ALSO apply category-weight emphasis"),
 ("","(e.g. union = labor x2.5); plus the current live grade for reference."),("",""),
 ("Model:","start at 100 -> deduct only for genuinely-bad categories (<48); worst issue"),
 ("","full, more severe issues compound, minor records fade; credit for verified-good;"),
 ("","thresholds A>=88 B>=70 C>=52 D>=36 F<36."),
 ("Tax:","EXCLUDED — legal tax behavior is a datapoint, not a demerit."),
 ("Directional:","political & DEI cut both ways (left helps a progressive, hurts a conservative;"),
 ("","pro-DEI helps a 'value DEI' user, hurts an 'avoid DEI' user)."),
 ("Priority issue:","weights the chosen category x2.5 (cares-most-about). '—' = neutral weights."),
 ("Caveat:","grades are only as good as the data — e.g. ExxonMobil's env record isn't in its"),
 ("","scored categories, so it reads cleaner than it should. Coverage gate + more sources fix this.")]
for i,(a,b) in enumerate(rows,1):
    ws3.cell(i,1,a).font=Font(bold=(i==1),size=12 if i==1 else 10); ws3.cell(i,2,b).font=Font(size=10)
for idx,g in zip(range(6,12),["A","B","C","D","F","?"]):
    ws3.cell(idx,1).fill=gfill(g); ws3.cell(idx,1).alignment=Alignment(horizontal="center"); ws3.cell(idx,1).font=Font(bold=True)
ws3.column_dimensions["A"].width=16; ws3.column_dimensions["B"].width=92

wb.save(OUT)
print("WROTE",OUT)
print("scenarios:",len(POL)*len(DEI)*len(ANI)*len(GUN)*len(PRI),"x",NA,"companies")
